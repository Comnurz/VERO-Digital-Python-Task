import argparse
import sys
from os.path import exists

import requests

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from datetime import datetime

from openpyxl.styles import PatternFill, Font


def get_color_code(hu):
    hu_not_older_than_three_months = "007500" # Green
    hu_not_older_than_twelve_months = "FFA500" # Orange
    hu_older_than_twelve_months = "b30000" # Red
    hu_time = datetime.strptime(hu, "%Y-%M-%d")
    now = datetime.now()
    if now + relativedelta(months=-3) < hu_time:
        return PatternFill(
            start_color=hu_not_older_than_three_months,
            end_color=hu_not_older_than_three_months,
            fill_type="solid"
        )
    elif now + relativedelta(months=-12) < hu_time:
        return PatternFill(
            start_color=hu_not_older_than_twelve_months,
            end_color=hu_not_older_than_twelve_months,
            fill_type="solid"
        )
    else:
        return PatternFill(
            start_color=hu_older_than_twelve_months,
            end_color=hu_older_than_twelve_months,
            fill_type="solid"
        )


def upload_csv(csv_path):
    print(f"Uploading {csv_path}...")
    res = requests.post("http://127.0.0.1:8000/upload", files={"file": open(csv_path, "rb")})
    res.raise_for_status()
    return res.json()


def generate_xlsx(vehicles, column_titles, is_colored=True):
    wb = Workbook()
    ws = wb.active
    ws.title = "vehicles"
    ws.append(column_titles)
    for item in vehicles:
        ws.append([item[title] for title in column_titles])
        if is_colored and 'hu' in column_titles:
            for cell in ws[ws.max_row]:
                cell.fill = get_color_code(item.get("hu"))

        if 'labelIds' in column_titles:
            cell = ws.cell(row=ws.max_row, column=column_titles.index('labelIds') + 1)
            if cell.value is not None and bool(cell.value):
                cell.font = Font(color=item.get('labelIds').split(',')[0])
    wb.save(f"vehicles_{datetime.now().isoformat()}.xlsx")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="CLI for generating excel from vehicle data")
    parser.add_argument("csv_path", type=str, help="Csv file path")
    parser.add_argument("-k", "--keys", type=str, help="Excel Keys")
    parser.add_argument("-c", "--colored", type=bool, default=True, help="Is colored?")
    args = parser.parse_args()
    csv_path = args.csv_path

    if not (csv_path.endswith(".csv") and exists(csv_path)):
        print("Invalid CSV path.")
        sys.exit(1)

    keys = {'gruppe', 'rnr'}
    if args.keys:
        keys = keys | set(args.keys.split(','))
    merged_list = upload_csv(csv_path)
    sorted_list = sorted(merged_list, key=lambda x: x['gruppe'])
    generate_xlsx(sorted_list, list(keys), args.colored)
